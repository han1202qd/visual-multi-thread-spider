from bs4 import BeautifulSoup
from selenium import webdriver
import time
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver import ChromeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.by import By
import re
import openpyxl
import os

def driver_init(language):
    desired_capabilities = DesiredCapabilities.CHROME
    desired_capabilities["pageLoadStrategy"] = "none"
    chrome_options = ChromeOptions()
    chrome_options.add_argument('--lang=%s'%language)
    #chrome_options.add_argument('--headless')                 # 无窗口模式
    chrome_options.add_argument('--disable-software-rasterizer')            #解决ERROR:gpu_init.cc(426) Passthrough is not supported, GL is swiftshader错误
    chrome_options.add_argument('--ignore-certificate-errors')              #忽略证书错误
    chrome_options.add_argument('-ignore-ssl-errors')
    chrome_options.add_argument('window-size=1920x3000')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-gpu')                            # 禁止硬件加速，避免严重占用cpu
    #chrome_options.add_argument('--disk-cache-dir=d:/chrome_cache')          # 自定义缓存目录
    #chrome_options.add_argument('–disk-cache-size= 20971520')           # 20M cache
    chrome_options.add_argument("disable-web-security")                     # 关闭安全策略
    chrome_options.add_experimental_option('prefs', {'profile.managed_default_content_settings.images': 1})     # 禁止图片加载
    chrome_options.add_argument('disable-infobars')                         # 隐藏"Chrome正在受到自动软件的控制
    chrome_options.add_experimental_option('excludeSwitches', ['enable-automation', 'enable-logging'])        # 设置开发者模式启动，该模式下webdriver属性为正常值,忽略无用日志，包括usb设备报错
    chrome_options.add_argument('user-agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36"')     # 模拟移动设备
    chrome_options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"       #手动指定本机电脑使用的浏览器位置    
    driver=webdriver.Chrome(executable_path=r"D:\GitHub\chromedriver\chromedriver.exe", options=chrome_options)  
    wait = WebDriverWait(driver, 30, 0.2)                                   # 返回驱动等待的变量
    driver.implicitly_wait(30)                                              # 隐性等待，最长等30秒
    driver.set_page_load_timeout(60)                                        #页面载入超时
    driver.set_script_timeout(60)                                           #脚本执行超时
    return wait,driver

def title_proc(title,num):
    title_1 = re.sub(r"\(.*?\)|\{.*?\}|\[.*?\]", "", title)           #将小括号，中括号，大括号及其中内容置空
    titles = re.findall(r"\w(?:[-'\w]*\w)?",title_1)                           #将标题分割为字母以及\'和连字符\-组成的单词
    i = 1
    return_title = ''
    for item in titles:
        if i <= num:
            return_title = return_title + ' ' + item
            i += 1
        else:
            break
    return return_title

def amazon_get_intro(wait,driver,url,title,author,delay):
    web = r'''https://www.amazon.com'''
    intro = ''
    driver.get(url)                                         #载入首页
    time.sleep(delay)
    try:
        wait.until(lambda x:x.find_element_by_id("asMain"))      #css选择器，返回结果存在跳出，异常报错   ec.presence_of_element_located((By.ID,"pageContent"))
        print('载入链接，开始输入数据:\t作者[%s]\t题名[%s]'%(author,title))
        
        #opt = driver.find_element_by_name("sort")
        #Select(opt).select_by_value("salesrank")
        driver.find_element_by_name("field-title").send_keys(title)
        driver.find_element_by_name("field-author").send_keys(author)
        print('开始检索书名[%s]'%title)
        driver.find_element_by_name("Adv-Srch-Books-Submit").click()          # driver自动点击搜索按钮，开始搜索

        time.sleep(delay)
        try:
            wait.until(lambda x:x.find_element_by_id("search"))      #css选择器，返回结果存在跳出，异常报错ec.presence_of_element_located((By.CLASS_NAME, "s-main-slot.s-result-list.s-search-results.sg-row"))
            list_driver = driver
            print('载入链接，开始处理检索结果')
            list_soup = BeautifulSoup(driver.page_source,'lxml')
            result = list_soup.find_all('div',class_='a-section a-spacing-small a-spacing-top-small')
            if result:
                print('\033[1;32m'+ result[0].text.replace('\n','') + '\033[0m' )
                search_list = list_soup.find_all('a',class_='a-link-normal a-text-normal')
                if search_list:
                    i = 0
                    for item in search_list:
                        i += 1
                        list_title = item.text
                        list_url = item['href']
                        print('开始载入第[%d]个检索结果[%s]'%(i, list_title))
                        one_url = web + list_url       
                        print('载入数据,链接为:[%s]'%one_url)
                        driver.get(one_url)                                         #转到内容页
                        time.sleep(delay)
                        try:
                            wait.until(lambda x:x.find_element_by_id("dp"))      #css选择器，返回结果存在跳出，异常报错ec.presence_of_element_located((By.ID, "dp"))
                            print('detail页载入完成')
                            iframe = driver.find_element_by_id("bookDesc_iframe")
                            driver.switch_to.frame(iframe)                              #转到iframe
                            print('载入iframe...')
                            time.sleep(delay)
                            try:
                                intro = wait.until(lambda x:x.find_element_by_id("iframeContent")).text      #css选择器，返回结果存在跳出，异常报错ec.presence_of_element_located((By.ID, "iframeContent"))
                                intro_length = len(re.findall(r"\w(?:[-'\w]*\w)?",intro))
                                kindle_text = re.findall(r"lecture Kindle",intro)
                                if intro_length > 15 and len(kindle_text) == 0:
                                    print('\033[1;32m' + intro + '\033[0m')
                                    break
                                else:
                                    print('获取简介过短[%d]或者包含kindle_text[%d]，继续尝试下一条检索结果...'%(intro_length,len(kindle_text)))
                            except Exception as e:
                                print('iframe没有加载完成/t',repr(e))
                        except  Exception as e:
                            print('detail页没有加载完成/t',repr(e))
                else:
                    print('没有发现题名为[%s]的检索结果列表'%title)
            else:
                print('[%s]没有检索到结果'%title)
        except Exception as e:
            print('检索列表页没有加载完成/t',repr(e))
    except Exception as e:
        print('高级检索页页没有加载完成/t',repr(e))
    return intro

def main(path,url):
    wb_dest = openpyxl.load_workbook(path)
    dest_sheet = wb_dest.worksheets[0]
    rows = dest_sheet.max_row
    success,fault = 0, 0
    start_time = time.time()
    circle_init = True
    for row in range(2,rows+1):
        print('\033[1;42m程序启动时间[%d],本次循环开始时间[%d],时间差为[%d]\033[0m'%(start_time, time.time(), time.time()-start_time))
        if time.time() - start_time > 600:
            if driver : driver.quit()
            clear_cache()
            start_time = time.time()
            circle_init = True
            
        if circle_init:
            language = 'en-us.UTF-8'
            wait, driver = driver_init(language)
            circle_init = False

        print('\033[1;31m' + "#" * 60 + '\033[0m')
        if dest_sheet.cell(row,4).value is not None:
            print('当前处理行[%d]'%row)
            book_name = dest_sheet.cell(row,4).value
            author_name = dest_sheet.cell(row,3).value
            intro, i, author, need_author = '', 10, '', True
            min_key = 3
            while len(intro)==0 and (i >= min_key):
                try:
                    print('当前标题单词数量为[%d]'% i )
                    title = title_proc(book_name, i)
                    i = i - 1
                    if author_name and need_author:
                        author = title_proc(author_name, 5)
                    intro = amazon_get_intro(wait, driver, url, title, author, 15)
                    intro_length = len(re.findall(r"\w(?:[-'\w]*\w)?",intro))
                    kindle_text = re.findall(r"lecture Kindle",intro)
                    if kindle_text or intro_length <10 :
                        intro = ''
                        print('获取简介过短[%d]或者包含kindle_text[%d],继续策略循环...'%(intro_length,len(kindle_text)))
                    if i == 2 and len(intro)==0 and need_author:        #如果指定作者检索一直检索不到内容，那么去掉作者检索
                        i = 10
                        need_author = False
                        author = ''
                        min_key = 4
                except Exception as e:
                    print('出现错误，继续循环策略..[%s]'%repr(e))
            if intro:
                success += 1
                dest_sheet.cell(row,13,value=intro)
                wb_dest.save(path)
                print('第[%d]行获取简介成功,目前共[%d]条成功记录'%(row,success))
            else:
                fault += 1
                print('第[%d]行没有获取简介,目前共[%d]条失败记录'%(row,fault))
            print('\033[1;31m' + "#" * 60 + '\033[0m')
            time.sleep(5)
    wb_dest.save(path)
    if driver: driver.quit()

def clear_cache():
    language = 'zh_CN.UTF-8'
    wait, driver = driver_init(language)
    url = 'chrome://settings/clearBrowserData'
    driver.get(url)
    time.sleep(5)
    clearButton = driver.execute_script("return document.querySelector('settings-ui').shadowRoot.querySelector('settings-main').shadowRoot.querySelector('settings-basic-page').shadowRoot.querySelector('settings-section > settings-privacy-page').shadowRoot.querySelector('settings-clear-browsing-data-dialog').shadowRoot.querySelector('#clearBrowsingDataDialog').querySelector('#clearBrowsingDataConfirm')")
    clearButton.click()
    time.sleep(10)
    driver.quit()



if __name__ == "__main__":
    os.system('')
    path = r'D:\GitHub\pyside\pdf_spider\france.xlsx'
    url = r'''https://www.amazon.com/advanced-search/books'''
    main(path,url)
     #clear_cache()

    
    



