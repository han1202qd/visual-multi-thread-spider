from PySide6.QtCore import *
from PySide6.QtGui import *
from PySide6.QtWidgets import *
from PySide6.QtTest import QTest
from ui_ChildWindow import Ui_Dialog
import openpyxl
import time
import os

class child_win(QDialog):
    """子窗口"""
    def __init__(self, table=None):
        super().__init__()
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)
        self.table = table
        self.filename = 'Save_Ebook.xlsx'
        self.timer = QTimer()
        #self.timer.timeout.connect(self.work)
        #self.timer.start(1000)
        self.timer.singleShot(1000, self.work)              #单次触发
    
    def work(self):
        rows = self.table.rowCount()
        self.ui.progressBar.setMaximum(rows)
        self.ui.progressBar.setValue(0)
        self.save()           
        self.close()
        self.deleteLater()
        
    def save(self):
        cols = self.table.columnCount()
        if os.path.exists(self.filename): 
            self.ui.label.setText("文件已存在，正在删除....")
            self.delay(1)
            os.remove(self.filename)
        self.ui.label.setText("正在生成文件....")
        self.delay(1)
        wb = openpyxl.Workbook()
        wb.create_sheet(index=1, title='失败记录')
        sheetlist = wb.worksheets
        sheetlist[0].title = '成功记录'
        for each in sheetlist:              #生成 两个工作表表头
            for col in range(cols):
                each.cell(row=1, column=col+1, value=self.table.horizontalHeaderItem(col).text())
        self.ui.label.setText("[成功记录][失败记录]表头已生成")
        self.delay(1)
        rows = self.table.rowCount()
        i = 0
        while i < rows:
            self.ui.label.setText("当前保存第[%d]记录,标题为[%s]"%(i,self.table.item(i,3).text()))
            self.delay(1)
            if self.table.item(i,0).background().color() == QColor('green'):                        #填充成功表数据
                sheet_rows = sheetlist[0].max_row
                for col in range(cols):
                    if self.table.item(i,col): sheetlist[0].cell(row=sheet_rows+1, column=col+1, value=self.table.item(i,col).text())

            if self.table.item(i,0).background().color() == QColor('red'):                        #填充失败表数据
                sheet_rows = sheetlist[1].max_row
                for col in range(cols):
                    if self.table.item(i,col): sheetlist[1].cell(row=sheet_rows+1, column=col+1, value=self.table.item(i,col).text())
            self.ui.label.setText("第[%d]记录保存完成"%i)
            self.delay(1)
            i += 1
            self.ui.progressBar.setValue(i)
            QCoreApplication.processEvents()
        self.ui.label.setText("正在保存文件....")
        self.delay(1)
        wb.save(self.filename)
        QMessageBox().information(self,'提示','信息已经保存完毕...', QMessageBox.Ok, QMessageBox.Ok)

    def delay(self,num):
        t = QElapsedTimer()
        t.start()
        while t.elapsed() < num * 100:    # 1/10秒为基础            
            QCoreApplication.processEvents()
            time.sleep(0.01)
        
