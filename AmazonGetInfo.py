# -*- coding:utf-8 -*-
from PySide6.QtCore import Qt, QObject, Signal, Slot, QEventLoop, QTimer, QThread, QTime, QMutex, QMutexLocker, QPoint, QDir
from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox, QWidget, QVBoxLayout, QTextEdit, QLineEdit, QLabel, QTableWidget, QTableWidgetItem, QGroupBox, QHBoxLayout, QListWidget, QToolTip, QHeaderView, QFileDialog
from PySide6.QtGui import QTextCursor, QCursor, QColor

import re
import os
import sys
import time
import requests
import openpyxl

from child import child_win
from ui_MainWindow import Ui_MainWindow
from GetIntro import WorkerThread, AmazonGet
from spy import spythread, spywork
from runtask import taskthread, taskwork
from SaveExcel import SaveExcelThread, SaveExcel
from globalvar import globalvar


# ====================================主窗口类=================================
class MainWindow(QMainWindow):
    """主窗口类"""
    PageAnalysisButtonSign = Signal(str)

    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.spinBox.setRange(1, 10)

        self.ui.tableWidget_5.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  # 平均分布各列

        self.ui.pushButton.clicked.connect(self.buttonclick)
        self.ui.pushButton_4.clicked.connect(self.getFiles)
        self.ui.button_switch_ip.clicked.connect(self.switch_ip)
        self.ui.tableWidget_3.itemClicked.connect(self.showMainTableRec)  # 点击显示表记录详细数据
        self.ui.pushButton_3.clicked.connect(self.clear_table)
        self.ui.pushButton_2.clicked.connect(self.open_dialog)
        self.begin_url = 'https://www.amazon.com/advanced-search/books'               # Amazon高级搜索
        self.ui.lineEdit.setText(self.begin_url)
        self.wids = dict()  # 动态生成控件集合
        self.threads = dict()  # 线程集合
        self.tasks = dict()  # 线程任务集合
        self.ui.tabWidget.setCurrentIndex(0)

    def creNewTab(self, num):
        tabtitle = u"线程" + str(num)
        tabname = u"tab" + str(num)
        textname = u"text" + str(num)
        linename = u"line" + str(num)
        labelname = u"lab" + str(num)
        listname = u"list" + str(num)

        curtab = QWidget()
        self.ui.tabWidget.addTab(curtab, tabtitle)
        curtab.setObjectName(tabname)

        curlayout = QVBoxLayout(curtab)

        curlab = QLabel(curtab)
        curlab.setObjectName(labelname)
        curlayout.addWidget(curlab)
        curlab.setMaximumHeight(20)
        curlab.setMinimumHeight(20)

        curline = QLineEdit(curtab, readOnly=True)
        curline.setObjectName(linename)
        #curline.setText(0)
        curlayout.addWidget(curline)

        curgroup = QGroupBox(curtab)
        curlayout.addWidget(curgroup)

        gblayout = QHBoxLayout(curgroup)

        curlist = QListWidget(curgroup)
        curlist.setObjectName(listname)
        curlist.setMaximumWidth(300)
        curlist.setMinimumWidth(300)
        gblayout.addWidget(curlist)

        curtext = QTextEdit(curgroup, readOnly=True)
        curtext.setObjectName(textname)
        gblayout.addWidget(curtext)

        self.wids[tabname] = curtab
        self.wids[textname] = curtext
        self.wids[linename] = curline
        self.wids[labelname] = curlab
        self.wids[listname] = curlist

    def tabswitch(self):
        cur_index = self.ui.tabWidget.currentIndex()
        self.ui.lineEdit.setText(str(cur_index))
        curtext = self.ui.tabWidget.currentWidget().findChild(QTextEdit)
        if curtext:
            curtext.append(curtext.objectName())

    def buttonclick(self):
        re_just = re.compile(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', re.IGNORECASE)
        target_url = self.ui.lineEdit.text()
        if re_just.match(target_url) is None:
            QMessageBox().critical(self, "错误", "非法网址，请修改初始网址")
            return

        if self.ui.tableWidget_3.rowCount() == 0:
            QMessageBox().critical(self, "错误", "还没有读入原始数据")
            return

        globalvar.url = self.ui.lineEdit.text()

        self.st = spythread(self, self.threads, self.wids)
        self.st.setObjectName('spythread')
        self.st.start()

        self.rt = taskthread(self, self.tasks, self.threads)
        self.rt.setObjectName('taskthread')
        self.threads['taskthread'] = self.rt
        self.rt.start()

        if self.ui.checkBox_3.isChecked():
            self.start_work()

        if self.ui.checkBox_4.isChecked():
            self.savethread = SaveExcelThread(self, self.ui.tableWidget_3)
            self.savethread.setObjectName('savetoexcel')
            self.threads['savetoexcel'] = self.savethread
            self.savethread.start()

    def build_thread(self, num):
        threadname = 'thread' + str(num)
        mt = WorkerThread(self, self.tasks)
        mt.setObjectName(threadname)
        self.threads[threadname] = mt
        mt.start()

    def start_work(self):
        num = self.ui.spinBox.value()
        for i in range(num):
            self.creNewTab(i)
            self.build_thread(i)

    def switch_ip(self):
        globalvar.must_change_ip = False
        self.ui.textEdit_4.append("已经手动切换VPN")
        self.ui.textEdit_4.append("当前使用IP：[%s]" % self.get_ip())

    @Slot(QTableWidgetItem)
    def showMainTableRec(self, it):  # 显示表格中的详细信息内容
        table = it.tableWidget()
        if table is None: return
        dt = self.ui.textEdit_2
        dt.clear()
        cols = table.columnCount()
        row = it.row()
        for col in range(cols):
            if table.item(row, col):
                showtxt = '[%s]:\t%s' % (table.horizontalHeaderItem(col).text(), table.item(row, col).text())
                dt.append(showtxt)

    def get_ip(self):
        header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36', }
        url = "http://httpbin.org/ip"
        proxies = {}
        try:
            response = requests.get(url, headers=header, proxies=proxies)
            temp = response.json()
            ip = temp.get('origin')
            return ip
        except Exception as e:
            return repr(e)

    def getFiles(self):
        msgwin = self.ui.textEdit_4
        table = self.ui.tableWidget_3
        dig = QFileDialog()
        dig.setFileMode(QFileDialog.AnyFile)
        # dig.setFilter(QDir.Files)
        dig.setNameFilter("Excel (*.xls *.xlsx )");
        msgwin.append('准备读取文件...')
        if dig.exec_():
            filenames = dig.selectedFiles()  # 接受选中文件的路径，默认为列表
            msgwin.append("打开文件[%s]" % filenames[0])
            file_name = os.path.basename(filenames[0])
            tcols = table.columnCount()
            wb = openpyxl.load_workbook(filenames[0])
            sheetlist = wb.worksheets
            rows = sheetlist[0].max_row
            
            cols = sheetlist[0].max_column
            table.setColumnCount( cols + 1 )                    #动态设置列名，需要先设置列数量
            column_name = []
            for col in range(1, cols + 1):                      # 生成表头
                column_name.append(sheetlist[0].cell(1, col).value)
            column_name.append('AMAZON_INTRO')                  # 表头增加一列
            table.setHorizontalHeaderLabels(column_name)

            if rows >= 2:
                cols = sheetlist[0].max_column                      # 重新计算列值
                for row in range(2, rows + 1):
                    trows = table.rowCount()
                    table.insertRow(trows)
                    rec_info = []
                    for col in range(1, cols + 1):
                        if col == 1:
                            table.setItem(trows, col - 1, QTableWidgetItem(file_name))
                            rec_info.append(file_name)
                        elif col == 2:
                            table.setItem(trows, col - 1, QTableWidgetItem(str(trows + 1)))
                            rec_info.append(str(trows + 1))
                        elif col == 3 or col == 4:
                            if sheetlist[0].cell(row, col).value is not None:
                                table.setItem(trows, col - 1, QTableWidgetItem(sheetlist[0].cell(row, col).value.replace('_x000D_',' ').replace('\n', ' ').replace('\r', ' ')))
                                rec_info.append(sheetlist[0].cell(row, col).value.replace('_x000D_', ' ').replace('\n', ' ').replace('\r', ' '))
                            else:
                                table.setItem(trows, col - 1, QTableWidgetItem(sheetlist[0].cell(row, col).value))
                                rec_info.append(sheetlist[0].cell(row, col).value)
                        else:
                            table.setItem(trows, col - 1, QTableWidgetItem(sheetlist[0].cell(row, col).value))
                            rec_info.append(sheetlist[0].cell(row, col).value)
                    globalvar.record_queue.put(rec_info)  # 放入队列
                    globalvar.record_count += 1
                msgwin.append("本次读取记录[%d]条" % (rows - 1))
            else:
                msgwin.append("文件[%s]无可用记录"%file_name)

    def clear_table(self):
        self.ui.tableWidget_3.clear()
        self.ui.tableWidget_3.setRowCount(0)
        self.ui.tableWidget_3.setColumnCount(0)
        globalvar.record_queue.queue.clear()
        globalvar.record_count = 0

    def open_dialog(self):
        self.childwin = child_win(self.ui.tableWidget_3)
        #self.childwin.show()                #非modal模式打开子窗口
        self.childwin.exec()                #modal模式打开子窗口


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.aboutToQuit.connect(app.deleteLater)
    window = MainWindow()
    window.showMaximized()
    sys.exit(app.exec_())

