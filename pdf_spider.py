# -*- coding:utf-8 -*-
from PySide6.QtCore import Qt, QObject, Signal, Slot, QEventLoop, QTimer, QThread, QTime, QMutex, QMutexLocker, QPoint, \
    QDir
from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox, QWidget, QVBoxLayout, QTextEdit, QProgressBar, \
    QLabel, QTableWidget, QTableWidgetItem, QGroupBox, QHBoxLayout, QListWidget, QToolTip, QHeaderView, QFileDialog
from PySide6.QtGui import QTextCursor, QCursor, QColor
from ui_MainWindow import Ui_MainWindow
from DocGet import download, WorkerThread
from PageAnalysis import getlist, page_info_thread
from spy import spythread, spywork
from runtask import taskthread, taskwork
from SaveExcel import SaveExcelThread, SaveExcel
from ProxyIP import ProxyThread, Get_Proxy
import re
import os
import sys
import time
import requests
import openpyxl
from globalvar import globalvar


# ====================================主窗口类=================================
class MainWindow(QMainWindow):
    """主窗口类"""
    PageAnalysisButtonSign = Signal(str)

    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.spinBox.setRange(1, 10)

        # self.ui.tableWidget.resizeColumnsToContents()
        # self.ui.tableWidget_2.resizeColumnsToContents()
        # self.ui.tableWidget_3.resizeColumnsToContents()
        # self.ui.tableWidget_4.resizeColumnsToContents()
        # self.ui.tableWidget_5.resizeColumnsToContents()
        self.ui.tableWidget_4.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  # 平均分布各列
        self.ui.tableWidget_5.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.ui.tableWidget.setColumnWidth(0,98)

        self.ui.pushButton.clicked.connect(self.buttonclick)
        self.ui.pushButton_2.clicked.connect(self.closetab)
        self.ui.pushButton_3.clicked.connect(self.pausePageAnalysis)
        self.ui.pushButton_4.clicked.connect(self.getFiles)
        # self.ui.tableWidget_3.itemEntered.connect(self.showTableItem)
        # self.ui.tableWidget_3.setMouseTracking(True)
        self.ui.tableWidget_3.itemClicked.connect(self.showMainTableRec)  # 两个表使用同一个槽函数
        # self.ui.tabWidget.currentChanged.connect(self.tabswitch)
        self.ui.checkBox.stateChanged.connect(self.checkbox_changed)
        # self.begin_url = 'https://www.ebooksgratuits.com/ebooks.php?option=search&id_auteur=0&id_categorie=1&id_genre=1&begin=0&offset=100'       # ebooksgratuits
        self.begin_url = 'https://www.gutenberg.org/ebooks/1'               # gutenberg
        self.ui.lineEdit.setText(self.begin_url)
        self.wids = dict()  # 动态生成控件集合
        self.threads = dict()  # 线程集合
        self.tasks = dict()  # 线程任务集合

    def creNewTab(self, num):
        tabtitle = u"线程" + str(num)
        tabname = u"tab" + str(num)
        textname = u"text" + str(num)
        progname = u"prog" + str(num)
        labelname = u"lab" + str(num)
        listname = u"list" + str(num)

        # self.ui.tabWidget.addTab(QWidget(), tabtitle)
        # self.ui.tabWidget.setCurrentIndex(self.ui.tabWidget.count() - 1)
        # curtab = self.ui.tabWidget.currentWidget()
        curtab = QWidget()
        self.ui.tabWidget.addTab(curtab, tabtitle)
        curtab.setObjectName(tabname)

        curlayout = QVBoxLayout(curtab)

        curlab = QLabel(curtab)
        curlab.setObjectName(labelname)
        curlayout.addWidget(curlab)
        curlab.setMaximumHeight(20)
        curlab.setMinimumHeight(20)

        curprog = QProgressBar(curtab)
        curprog.setObjectName(progname)
        curprog.setValue(0)
        curlayout.addWidget(curprog)

        curgroup = QGroupBox(curtab)
        curlayout.addWidget(curgroup)

        gblayout = QHBoxLayout(curgroup)

        curlist = QListWidget(curgroup)
        curlist.setObjectName(listname)
        curlist.setMaximumWidth(300)
        curlist.setMinimumWidth(300)
        gblayout.addWidget(curlist)

        curtext = QTextEdit(curgroup, readOnly=True)
        curtext.setObjectName(textname)
        gblayout.addWidget(curtext)

        self.wids[tabname] = curtab
        self.wids[textname] = curtext
        self.wids[progname] = curprog
        self.wids[labelname] = curlab
        self.wids[listname] = curlist

    def tabswitch(self):
        cur_index = self.ui.tabWidget.currentIndex()
        self.ui.lineEdit.setText(str(cur_index))
        curtext = self.ui.tabWidget.currentWidget().findChild(QTextEdit)
        if curtext:
            curtext.append(curtext.objectName())

    def buttonclick(self):
        re_just = re.compile(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', re.IGNORECASE)
        target_url = self.ui.lineEdit.text()
        if re_just.match(target_url) is None:
            QMessageBox().critical(self, "错误", "非法网址，请修改初始网址")
            return

        if self.ui.checkBox.isChecked():
            globalvar.use_proxy = True
        else:
            globalvar.use_proxy = False
            self.ui.textEdit_3.append("当前不使用代理服务")
            self.ui.textEdit_3.append("当前使用IP：[%s]" % self.get_ip())

        self.st = spythread(self, self.threads, self.wids)
        self.st.setObjectName('spythread')
        self.st.start()

        if globalvar.use_proxy:
            self.GetIpt = ProxyThread(self)
            self.GetIpt.setObjectName('ProxyIP')
            self.threads['ProxyIP'] = self.GetIpt
            self.GetIpt.start()

        if self.ui.checkBox_2.isChecked():
            self.pt = page_info_thread(self, target_url)
            self.pt.setObjectName('pageAnalysis')
            self.threads['pageAnalysis'] = self.pt
            self.pt.start()

        self.rt = taskthread(self, self.tasks, self.threads)
        self.rt.setObjectName('taskthread')
        self.threads['taskthread'] = self.rt
        self.rt.start()

        if self.ui.checkBox_3.isChecked():
            self.start_work()

        if self.ui.checkBox_4.isChecked():
            self.savethread = SaveExcelThread(self, self.ui.tableWidget_3)
            self.savethread.setObjectName('savetoexcel')
            self.threads['savetoexcel'] = self.savethread
            self.savethread.start()

    def build_thread(self, num):
        threadname = 'thread' + str(num)
        mt = WorkerThread(self, self.tasks)
        mt.setObjectName(threadname)
        self.threads[threadname] = mt
        mt.start()

    def start_work(self):
        num = self.ui.spinBox.value()
        for i in range(num):
            self.creNewTab(i)
            self.build_thread(i)

    def closetab(self):
        # self.detroytab(self.wids['tab1'])
        # for key in self.tasks:
        #    self.ui.textEdit_4.append('[%d][%d][%d]'%(self.tasks[key][0],self.tasks[key][1],self.tasks[key][2]))
        globalvar.must_change_ip = False
        self.ui.textEdit_3.append("已经手动切换VPN")
        self.ui.textEdit_3.append("当前使用IP：[%s]" % self.get_ip())

    # def detroytab(self,child):
    #    child.deleteLater()

    def pausePageAnalysis(self):  # 暂停页面分析进程
        if self.threads.get('pageAnalysis') and self.threads['pageAnalysis'].isRunning():
            if self.ui.pushButton_3.text() == '暂停页面分析':
                self.PageAnalysisButtonSign.emit('pause')
                self.ui.pushButton_3.setText('继续页面分析')
            else:
                self.PageAnalysisButtonSign.emit('continue')
                self.ui.pushButton_3.setText('暂停页面分析')

    @Slot(QTableWidgetItem)
    def showTableItem(self, it):  # 下载详细记录列表的TOOLTIP
        if it == None:
            return
        curtable = self.ui.tableWidget_3
        QToolTip.showText(QCursor.pos(), it.text())
        # curtable = self.ui.tableWidget_3
        # QToolTip.hideText()
        # r = curtable.visualItemRect(it)
        # p = curtable.viewport().mapToGlobal(QPoint(r.center().x(), r.top()))
        # QToolTip.showText(p, it.text(), curtable, curtable.rect(),50000 )           #此处无法定义显示时间，还是有问题

    @Slot(QTableWidgetItem)
    def showMainTableRec(self, it):  # 显示表格中的详细信息内容
        table = it.tableWidget()
        if table is None: return
        dt = self.ui.textEdit_2
        dt.clear()
        cols = table.columnCount()
        row = it.row()
        for col in range(cols):
            if table.item(row, col):
                showtxt = '[%s]:\t%s' % (table.horizontalHeaderItem(col).text(), table.item(row, col).text())
                dt.append(showtxt)

    @Slot(list)
    def upd_iptable_status(self, ilist):
        table = self.ui.tableWidget_4
        items = table.findItems(ilist[1], Qt.MatchExactly)
        if items:
            for item in items:
                if ilist[0] == 'True':
                    item.setBackground(QColor('green'))
                    item.setForeground(QColor('white'))
                else:
                    item.setBackground(QColor('red'))
                    item.setForeground(QColor('white'))

    def get_ip(self):
        header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36', }
        url = "http://httpbin.org/ip"
        proxies = {}
        try:
            response = requests.get(url, headers=header, proxies=proxies)
            temp = response.json()
            ip = temp.get('origin')
            return ip
        except Exception as e:
            return repr(e)

    def checkbox_changed(self):
        if self.threads != {}:
            if self.ui.checkBox.checkState() == Qt.Checked:
                if self.threads.get('ProxyIP') is None:
                    globalvar.ip_end = False
                    globalvar.use_proxy = True
                    self.GetIpt = ProxyThread(self)
                    self.GetIpt.setObjectName('ProxyIP')
                    self.threads['ProxyIP'] = self.GetIpt
                    self.GetIpt.start()
            elif self.ui.checkBox.checkState() == Qt.Unchecked:
                if self.threads.get('ProxyIP'):
                    globalvar.use_proxy = False
                    globalvar.ip_end = True
                    # self.threads['ProxyIP'].quit()
                    # self.threads['ProxyIP'].wait()

    def getFiles(self):
        if len(self.threads) == 0:
            if self.ui.checkBox.isChecked():
                globalvar.use_proxy = True
            else:
                globalvar.use_proxy = False
                self.ui.textEdit_3.append("当前不使用代理服务")
                self.ui.textEdit_3.append("当前使用IP：[%s]" % self.get_ip())

            self.st = spythread(self, self.threads, self.wids)
            self.st.setObjectName('spythread')
            self.st.start()

            if globalvar.use_proxy:
                self.GetIpt = ProxyThread(self)
                self.GetIpt.setObjectName('ProxyIP')
                self.threads['ProxyIP'] = self.GetIpt
                self.GetIpt.start()

            self.rt = taskthread(self, self.tasks, self.threads)
            self.rt.setObjectName('taskthread')
            self.threads['taskthread'] = self.rt
            self.rt.start()

            self.start_work()

            self.savethread = SaveExcelThread(self, self.ui.tableWidget_3)
            self.savethread.setObjectName('savetoexcel')
            self.threads['savetoexcel'] = self.savethread
            self.savethread.start()

        msgwin = self.ui.textEdit
        table = self.ui.tableWidget_3
        dig = QFileDialog()
        dig.setFileMode(QFileDialog.AnyFile)
        # dig.setFilter(QDir.Files)
        dig.setNameFilter("Excel (*.xls *.xlsx )");
        if dig.exec_():
            filenames = dig.selectedFiles()  # 接受选中文件的路径，默认为列表
            msgwin.append("打开文件[%s]" % filenames[0])
            file_name = filenames[0].split("/")[-1].split(".")[0]
            tcols = table.columnCount()
            wb = openpyxl.load_workbook(filenames[0])
            sheetlist = wb.worksheets
            if sheetlist[1].title == "失败记录":
                rows = sheetlist[1].max_row
                if rows >= 2:
                    cols = sheetlist[1].max_column
                    for row in range(2, rows + 1):
                        trows = table.rowCount()
                        table.insertRow(trows)
                        rec_info = []
                        for col in range(1, cols + 1):
                            if col == 1:
                                table.setItem(trows, col - 1, QTableWidgetItem(file_name))
                                rec_info.append(file_name)
                            elif col == 2:
                                table.setItem(trows, col - 1, QTableWidgetItem(str(trows + 1)))
                                rec_info.append(str(trows + 1))
                            else:
                                table.setItem(trows, col - 1, QTableWidgetItem(sheetlist[1].cell(row, col).value))
                                rec_info.append(sheetlist[1].cell(row, col).value)
                        globalvar.url_queue.put(rec_info)  # 放入队列
                        globalvar.url_count += 1
                    msgwin.append("本次读取失败记录[%d]条" % (rows - 1))
                else:
                    msgwin.append("本次读取失败[0]条")
            else:
                msgwin.append("没有发现失败记录")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.aboutToQuit.connect(app.deleteLater)
    window = MainWindow()
    window.showMaximized()
    sys.exit(app.exec_())
