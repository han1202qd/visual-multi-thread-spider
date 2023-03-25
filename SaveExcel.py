from PySide6.QtCore import QObject, Signal, Slot, QEventLoop, QTimer, QThread, QMutex, QMutexLocker, QPoint
from PySide6.QtWidgets import QTableWidget, QTableWidgetItem
import openpyxl
import re
import os
import sys
import time
from globalvar import globalvar

#====================================EXCEL工作线程类=================================
class SaveExcelThread(QThread):
    """EXCEL工作线程"""
    def __init__(self, parent=None, table=None):
        QThread.__init__(self, parent)
        self.form = parent
        self.savework = SaveExcel(self,table)


    def run(self):
        self.savework.start_work()
        self.exec_()

    @Slot(str)
    def upd_saveinfo(self,text):
        self.form.ui.textEdit_4.append(text)

#====================================保存EXCEL文件=================================
class SaveExcel(QObject):
    """保存EXCEL文件"""
    info_sign = Signal(str)

    def __init__(self, parent=None, table=None):
        super(self.__class__, self).__init__(parent)
        self.table = table
        self.filename = 'Ebook.xlsx'
        self.firstRun = True
        self.info_sign.connect(parent.upd_saveinfo)

    def start_work(self):
        self.timer = QTimer()
        self.timer.timeout.connect(self.SaveTable)
        self.timer.start(120000)

    def SaveTable(self):
        #try:
        self.printstr("EXCEL保存正在启动[%s]"%time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))
        cols = self.table.columnCount()
        if os.path.isfile(self.filename) and (not self.firstRun):
            wb = openpyxl.load_workbook(self.filename)
            sheetlist = wb.worksheets

            rows = sheetlist[0].max_row
            j = 0
            while not globalvar.success_queue.empty():                                                #先填充成功表数据                   
                irow = int( globalvar.success_queue.get() ) - 1
                self.printstr("成功记录：[%d]"%irow)
                j += 1
                try:
                    for col in range(cols):
                        if self.table.item(irow,col): sheetlist[0].cell(row=rows+j, column=col+1, value=self.table.item(irow,col).text())
                except Exception as e:
                    self.printstr('存储记录[%d]时出错\t%s'%(irow,repr(e)))
            self.printstr("本次[成功记录]保存[%d]条"%j)
            globalvar.excel_success_count = rows + j            #成功记录计数

            rows = sheetlist[1].max_row
            j = 0
            while not globalvar.fault_queue.empty():                                                    #然后填充失败表数据
                irow = int( globalvar.fault_queue.get() ) - 1
                self.printstr("失败记录：[%d]"%irow)
                j += 1
                try:
                    for col in range(cols):
                        if self.table.item(irow,col): sheetlist[1].cell(row=rows+j, column=col+1, value=self.table.item(irow,col).text())
                except Exception as e:
                    self.printstr('存出记录[%d]时出错\t%s'%(irow,repr(e)))
            self.printstr("本次[失败记录]保存[%d]条"%j)
            globalvar.excel_fault_count = rows + j            #失败记录计数
        else:
            if os.path.exists(self.filename): os.remove(self.filename)
            wb = openpyxl.Workbook()
            wb.create_sheet(index=1, title='失败记录')
            sheetlist = wb.worksheets
            sheetlist[0].title = '成功记录'
            for each in sheetlist:              #生成 两个工作表表头
                for col in range(cols):
                    each.cell(row=1, column=col+1, value=self.table.horizontalHeaderItem(col).text())
            self.printstr("[成功记录][失败记录]表头已生成")

            j = 0
            while not globalvar.success_queue.empty():                                                #先填充成功表数据                   
                irow = int( globalvar.success_queue.get() ) - 1
                self.printstr("成功记录：[%d]"%irow)
                j += 1
                try:
                    for col in range(cols):
                        if self.table.item(irow,col): sheetlist[0].cell(row=j+1, column=col+1, value=self.table.item(irow,col).text())
                except Exception as e:
                    self.printstr('存储记录[%d]时出错\t%s'%(irow,repr(e)))
            self.printstr("本次[成功记录]保存[%d]条"%j)
            globalvar.excel_success_count = j            #成功记录计数

            j = 0
            while not globalvar.fault_queue.empty():                                                    #然后填充失败表数据
                irow = int( globalvar.fault_queue.get() ) - 1
                self.printstr("失败记录：[%d]"%irow)
                j += 1
                try:
                    for col in range(cols):
                        if self.table.item(irow,col): sheetlist[1].cell(row=j+1, column=col+1, value=self.table.item(irow,col).text())
                except Exception as e:
                    self.printstr('存储记录[%d]时出错\t%s'%(irow,repr(e)))
            self.printstr("本次[失败记录]保存[%d]条"%j)
            globalvar.excel_fault_count = j            #失败记录计数
        self.firstRun = False
        wb.save(self.filename)
        self.printstr("本次文件保存完成")
        #except Exception as e:
        #    self.printstr('EXCEL文件被另一进程占用，下个循环重新尝试写入数据\t%s'%repr(e))

    def printstr(self,text):
        self.info_sign.emit(text)

            

            
            

        

